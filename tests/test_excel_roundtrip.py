"""
Tests for the Excel round-trip workflow:
  download labels -> export to .xlsx -> user edits New_Label -> load -> upload
"""
import tempfile
from pathlib import Path

import openpyxl
import pytest

from src.agents.file_handler.excel_handler import ExcelHandler
from src.agents.file_handler.schema import FileData, PortData


@pytest.fixture
def handler():
    return ExcelHandler()


@pytest.fixture
def sample_data():
    """Simulates data downloaded from a 4x4 router."""
    return FileData(ports=[
        PortData(port=1, type="INPUT", current_label="Camera 1"),
        PortData(port=2, type="INPUT", current_label="Camera 2"),
        PortData(port=3, type="INPUT", current_label="Laptop"),
        PortData(port=4, type="INPUT", current_label="Replay"),
        PortData(port=1, type="OUTPUT", current_label="Program"),
        PortData(port=2, type="OUTPUT", current_label="Preview"),
        PortData(port=3, type="OUTPUT", current_label="Record"),
        PortData(port=4, type="OUTPUT", current_label="Stream"),
    ])


# ── Basic write / read round-trip ─────────────────────────────────

def test_roundtrip_preserves_data(handler, sample_data):
    """Write then read should preserve all port data."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = Path(f.name)

    handler.write_excel(path, sample_data)
    loaded = handler.read_excel(path)

    assert len(loaded.ports) == len(sample_data.ports)
    for orig, loaded_port in zip(
        sorted(sample_data.ports, key=lambda p: (p.type, p.port)),
        sorted(loaded.ports, key=lambda p: (p.type, p.port)),
    ):
        assert loaded_port.port == orig.port
        assert loaded_port.type == orig.type
        assert loaded_port.current_label == orig.current_label


def test_roundtrip_no_changes_by_default(handler, sample_data):
    """Exported file should have no pending changes (new_label=None)."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = Path(f.name)

    handler.write_excel(path, sample_data)
    loaded = handler.read_excel(path)

    for port in loaded.ports:
        assert port.new_label is None, (
            f"Port {port.port} ({port.type}) should have no pending change"
        )


# ── User edits the New_Label column ──────────────────────────────

def test_user_edits_are_detected(handler, sample_data):
    """Simulate a user typing new names in the yellow column."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = Path(f.name)

    handler.write_excel(path, sample_data)

    # Open the file and simulate user edits in the New_Label column (D)
    wb = openpyxl.load_workbook(path)
    ws = wb[ExcelHandler.WORKSHEET_NAME]

    # Find the row for INPUT port 1 and write a new label
    for row in ws.iter_rows(min_row=2, max_col=4):
        port_val = row[0].value
        type_val = row[1].value
        if port_val == 1 and type_val == "INPUT":
            row[3].value = "Main Camera"
            break

    wb.save(path)
    wb.close()

    # Re-load and verify
    loaded = handler.read_excel(path)
    input_1 = next(p for p in loaded.ports if p.port == 1 and p.type == "INPUT")
    assert input_1.new_label == "Main Camera"
    assert input_1.current_label == "Camera 1"  # unchanged

    # Other ports should still have no changes
    others = [p for p in loaded.ports if not (p.port == 1 and p.type == "INPUT")]
    for port in others:
        assert port.new_label is None


def test_blank_new_label_means_no_change(handler, sample_data):
    """Whitespace-only or empty New_Label should be treated as no change."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = Path(f.name)

    handler.write_excel(path, sample_data)

    wb = openpyxl.load_workbook(path)
    ws = wb[ExcelHandler.WORKSHEET_NAME]

    # Set some cells to whitespace / empty string
    for row in ws.iter_rows(min_row=2, max_col=4):
        port_val = row[0].value
        if port_val == 2:
            row[3].value = "   "  # whitespace only
        elif port_val == 3:
            row[3].value = ""     # empty string

    wb.save(path)
    wb.close()

    loaded = handler.read_excel(path)
    for port in loaded.ports:
        if port.port in (2, 3):
            assert port.new_label is None, (
                f"Port {port.port} with blank New_Label should be None"
            )


# ── Section separator handling ────────────────────────────────────

def test_section_separator_is_skipped_on_read(handler, sample_data):
    """The OUTPUTS separator row must not produce a PortData entry."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = Path(f.name)

    handler.write_excel(path, sample_data)
    loaded = handler.read_excel(path)

    # Should have same count as original – no separator artefact
    assert len(loaded.ports) == len(sample_data.ports)

    types = {p.type for p in loaded.ports}
    assert types == {"INPUT", "OUTPUT"}


# ── Instructions sheet ────────────────────────────────────────────

def test_instructions_sheet_exists(handler, sample_data):
    """The exported workbook should contain an Instructions sheet."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = Path(f.name)

    handler.write_excel(path, sample_data)

    wb = openpyxl.load_workbook(path)
    assert "Instructions" in wb.sheetnames
    assert ExcelHandler.WORKSHEET_NAME in wb.sheetnames
    wb.close()


def test_data_sheet_is_first_tab(handler, sample_data):
    """KUMO_Labels should be the first sheet (active on open)."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = Path(f.name)

    handler.write_excel(path, sample_data)

    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames[0] == ExcelHandler.WORKSHEET_NAME
    wb.close()


# ── Template ──────────────────────────────────────────────────────

def test_template_creates_64_ports(handler):
    """Template should produce 32 inputs + 32 outputs."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = Path(f.name)

    handler.write_excel(path, FileData(ports=[]), create_template=True)
    loaded = handler.read_excel(path)

    assert len(loaded.get_inputs()) == 32
    assert len(loaded.get_outputs()) == 32


# ── Protection ────────────────────────────────────────────────────

def test_sheet_protection_enabled(handler, sample_data):
    """Sheet protection should be on to guard read-only columns."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = Path(f.name)

    handler.write_excel(path, sample_data)

    wb = openpyxl.load_workbook(path)
    ws = wb[ExcelHandler.WORKSHEET_NAME]
    assert ws.protection.sheet is True
    wb.close()


def test_new_label_column_is_unlocked(handler, sample_data):
    """The New_Label (D) cells should be unlocked for editing."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = Path(f.name)

    handler.write_excel(path, sample_data)

    wb = openpyxl.load_workbook(path)
    ws = wb[ExcelHandler.WORKSHEET_NAME]

    # Check a few data rows in column D
    for row_idx in range(2, 6):
        cell = ws.cell(row=row_idx, column=4)
        assert cell.protection.locked is False, (
            f"Row {row_idx} New_Label cell should be unlocked"
        )
    wb.close()
