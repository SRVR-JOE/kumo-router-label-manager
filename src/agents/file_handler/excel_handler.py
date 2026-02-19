"""
Excel file handler using openpyxl.
"""
from pathlib import Path
from typing import Optional
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from .schema import FileData, PortData


class ExcelHandler:
    """Handler for Excel file operations."""

    WORKSHEET_NAME = "KUMO_Labels"
    HEADERS = ["Port", "Type", "Current_Label", "New_Label", "Notes"]

    def __init__(self):
        """Initialize the Excel handler."""
        pass

    def read_excel(self, file_path: Path) -> FileData:
        """
        Read Excel file and parse KUMO_Labels worksheet.

        Args:
            file_path: Path to Excel file

        Returns:
            FileData object with parsed data

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If worksheet not found or data is invalid
        """
        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            raise ValueError(f"Failed to load Excel file: {e}")

        if self.WORKSHEET_NAME not in workbook.sheetnames:
            raise ValueError(
                f"Worksheet '{self.WORKSHEET_NAME}' not found. "
                f"Available sheets: {workbook.sheetnames}"
            )

        worksheet = workbook[self.WORKSHEET_NAME]
        ports = []

        # Start from row 2 (skip header)
        for row_idx in range(2, worksheet.max_row + 1):
            row = worksheet[row_idx]

            # Skip empty rows
            if all(cell.value is None for cell in row):
                continue

            try:
                port_value = row[0].value
                if port_value is None:
                    raise ValueError("Port number is required")

                current_label = str(row[2].value).strip() if row[2].value else ""
                new_label = str(row[3].value).strip() if row[3].value else None
                notes = str(row[4].value).strip() if len(row) > 4 and row[4].value else ""

                port_data = PortData(
                    port=int(port_value),
                    type=str(row[1].value).strip() if row[1].value else "INPUT",
                    current_label=current_label,
                    new_label=new_label,
                    notes=notes,
                )
                ports.append(port_data)
            except (ValueError, TypeError) as e:
                raise ValueError(f"Invalid data in row {row_idx}: {e}")

        workbook.close()
        return FileData(ports=ports)

    def write_excel(
        self,
        file_path: Path,
        data: FileData,
        create_template: bool = False
    ) -> None:
        """
        Write data to Excel file with formatting.

        Args:
            file_path: Path to save Excel file
            data: FileData object to write
            create_template: If True, create a 64-row template

        Raises:
            ValueError: If data is invalid
        """
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self.WORKSHEET_NAME

        # Write headers
        for col_idx, header in enumerate(self.HEADERS, start=1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data or create template
        if create_template:
            # Create 64-row template: 32 inputs + 32 outputs
            for i in range(1, 33):
                self._write_row(worksheet, i + 1, i, "INPUT")
            for i in range(1, 33):
                self._write_row(worksheet, i + 33, i, "OUTPUT")
        else:
            # Write actual data
            for row_idx, port_data in enumerate(data.ports, start=2):
                self._write_row(
                    worksheet,
                    row_idx,
                    port_data.port,
                    port_data.type,
                    port_data.current_label,
                    port_data.new_label or "",
                    port_data.notes,
                )

        # Add data validation for Type column (column B)
        type_validation = DataValidation(
            type="list",
            formula1='"INPUT,OUTPUT"',
            allow_blank=False,
            showDropDown=True,
        )
        type_validation.error = "Please select INPUT or OUTPUT"
        type_validation.errorTitle = "Invalid Type"

        # Apply to all data rows
        max_row = 65 if create_template else len(data.ports) + 1
        type_validation.add(f"B2:B{max_row}")
        worksheet.add_data_validation(type_validation)

        # Auto-fit columns
        self._autofit_columns(worksheet)

        # Freeze header row
        worksheet.freeze_panes = "A2"

        # Save workbook
        try:
            workbook.save(file_path)
        except Exception as e:
            raise ValueError(f"Failed to save Excel file: {e}")
        finally:
            workbook.close()

    def _write_row(
        self,
        worksheet,
        row_idx: int,
        port: int,
        port_type: str,
        current_label: str = "",
        new_label: str = "",
        notes: str = "",
    ) -> None:
        """Write a single row to the worksheet."""
        worksheet.cell(row=row_idx, column=1, value=port)
        worksheet.cell(row=row_idx, column=2, value=port_type)
        worksheet.cell(row=row_idx, column=3, value=current_label)
        worksheet.cell(row=row_idx, column=4, value=new_label)
        worksheet.cell(row=row_idx, column=5, value=notes)

        # Center align port and type columns
        worksheet.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")
        worksheet.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")

    def _autofit_columns(self, worksheet) -> None:
        """Auto-fit column widths based on content."""
        for column_cells in worksheet.columns:
            max_length = 0
            column = column_cells[0].column_letter

            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except (ValueError, TypeError, AttributeError):
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            worksheet.column_dimensions[column].width = max(adjusted_width, 10)
